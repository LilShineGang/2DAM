# Programa 3: Infraestructuras 5G - Medición de velocidad

# Se importa desde la librería openpyxl dos clases/funciones:
# Workbook: para crear un nuevo libro de Excel (.xlsx).
# load_workbook: para abrir...leer un archivo .xlsx ya existente.
from openpyxl import Workbook, load_workbook

# Crear archivo con datos de pruebas
wb = Workbook()
ws = wb.active  # se toma la hoja activa por defecto
ws.title = "Pruebas5G"

ws.append(["Zona", "Velocidad (Mbps)"])
ws.append(["Centro", 1200])
ws.append(["Periferia", 850])
ws.append(["Aeropuerto", 950])

wb.save("redes5G.xlsx")

# Leer archivo generado
wb2 = load_workbook("redes5G.xlsx")
ws2 = wb2.active
print("=== Velocidades registradas ===")
# iter_rows devuelve tuplas con los valores de cada celda
for fila in ws2.iter_rows(values_only=True):
    print(fila)ut(
        "\n2: ¿Cuáles son las causas de la contaminación de nuestros acuíferos?\n"
    )
    r3 = input("\n3: ¿Cuál es la solución propuesta?\n")
    r4 = input("\n4: ¿Qué es la sopa verde?\n")
    r5 = input(
        "\n5: ¿Qué actividades realizadas por la agroindustria son causantes de este colapso según David Verdiell?\n"
    )

    guardar(r1, r2, r3, r4, r5)


if __name__ == "__main__":
    main()
